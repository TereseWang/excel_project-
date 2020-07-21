import xlrd
import click
import openpyxl
from time import strptime
import copy
import datetime


@click.group()
def reformat():
    "Reformat Data files"

@reformat.command('reformat')
@click.option('--file', default=None,
    help="Only files for AUM&Data Metrics")
def reformat_file(file):
    "Entry point for getting the file"
    """Type: "python reformat.py reformat --file "filename"" to run"""
    book = xlrd.open_workbook(file)
    names = book.sheet_names()
    sheets = {}
    for name in names:
        sheets[name] = copy.copy(book.sheet_by_name(name))

    rbook = openpyxl.load_workbook('reformat.xlsx')

    cal_sheet = rbook['Calculated AUM&Account Metrics']
    cal_reformat_sheet = reformat_calculated(sheets, cal_sheet)

    rbook.save('reformat.xlsx')
    return

def search_row(name,colindex,sheet):
    value = None
    for row in range(sheet.nrows):
        if sheet.cell_value(row, colindex) == name:
            value = sheet.row_values(row)
            break
        else:
            pass
    return value

def search_column(name, rowindex, sheet):
    value = None
    for column in range(sheet.ncols):
        if sheet.cell_value(rowindex, column) == name:
            value = sheet.col_values(column)
            break
        else:
            pass
    return value

def find_thing_rown(name, sheet):
    value = None
    for row in sheet.rows:
        for cell in row:
            if str(cell.value) == name:
                value = cell.row
                break
        else:
            continue
        break
    return value

def current_time(sheets):
    month = search_row('Current Month', 1, sheets['input'])[2]
    year = search_row('Current FY', 1, sheets['input'])[2]
    month_num = strptime(str(month),'%B').tm_mon
    return year[0:4] + "-" + "{0:0=2d}".format(month_num) + "-01 00:00:00"

def next_month(time):
    year = int(time[0:4])
    month = int(time[5:7])
    if month == 12:
        month = 1
        year = year + 1
    else:
        month = month + 1
    return "{0}-{1}-01 00:00:00".format(str(year), "{0:0=2d}".format(month))

def end_quarter_month(time):
    end_time = None
    if int(time[5:7]) < 10 :
        end_time = "{0}-10-01 00:00:00".format(time[0:4])
    else:
        end_time = "{0}-10-01 00:00:00".format(str(int(time[0:4]) + 1))
    return end_time

def date_format (time, format):
    return datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S").strftime(format)

def update_col(lov, sr, er, ci, sheet):
    index = 0
    for row in range(sr, er):
        sheet.cell(row, ci).value = lov[index]
        index = index + 1
    return sheet

def reformat_calculated(sheets, sheet):
    "Reformat calculated AUM&Account Metrics sheet"
    time = current_time(sheets)
    current_rown = find_thing_rown(time, sheet)
    future = next_month(time)
    future_rown = current_rown + 12
    end = end_quarter_month(future)
    currentv = search_column(date_format(time, "%B"), 1, sheets['Data'])[2:]
    futurev = search_column(date_format(future, "%B"), 1, sheets['Data'])[2:]
    forecastv = search_column('Full Year Forecast', 1, sheets['Data'])[2:]
    list_of_actuals = ['Actual'] * 11

    if future[5:7] == end[5:7]:
        update_col(list_of_actuals, current_rown, current_rown+10,3, sheet)
        update_col(currentv, current_rown, current_rown+10, 5, sheet)
        update_col(futurev, future_rown, future_rown+10,5, sheet)

    else:
        update_col(list_of_actuals, current_rown, current_rown+11,3, sheet)
        update_col(currentv, current_rown, current_rown+11, 5, sheet)

        sheet.move_range("A{0}:E{1}".format(future_rown, future_rown+10), rows=+12, cols=0)
        update_col(forecastv, future_rown+12, future_rown+23, 5, sheet)
        index = 0
        for row in range(future_rown+12, future_rown + 23):
            if int(futurev[index]) == 0:
                sheet.cell(row, 5).value = 'N/A'
            else:
                sheet.cell(row, 5).value = futurev[index]
            index = index + 1

        list_of_forecasts = ['Forecast'] * 11
        update_col(list_of_forecasts, future_rown, future_rown+11, 3, sheet)
        list_of_future_dates = [date_format(future, "%m-%d-%Y")] * 11
        update_col(list_of_future_dates, future_rown, future_rown+11, 4, sheet)
        index1 = 0
        for row in range(future_rown, future_rown + 11):
            sheet.cell(row, 1).value = sheet.cell(row - 12, 1).value
            sheet.cell(row, 2).value = sheet.cell(row - 12, 2).value
            if int(futurev[index1]) == 0:
                sheet.cell(row, 5).value = 'N/A'
                index1 = index1 + 1
            else:
                sheet.cell(row, 5).value = futurev[index1]
                index1 = index1 + 1
    return sheet

if __name__ == '__main__':
    reformat()
